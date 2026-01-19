"""
Export script to create Excel file with MISMATCHED records.
Includes Type-Safety and adds ACTUAL vs PREDICTED columns for easier debugging.
"""

import sqlite3
import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tqdm import tqdm
from typing import Optional

# =============================================================================
# CONFIGURATION
# =============================================================================

GROUND_TRUTH_FILE = "data_check.csv"  # Ensure this file exists!
PREDICTED_FILE = "analysis_output/classified_data_active_users.csv"
COLUMNS_TO_VALIDATE = ["ir_user"]  # Can add "ir_user", "fx_user"

# =============================================================================
# 1. IDENTIFY MISMATCHES & RETRIEVE VALUES
# =============================================================================


def get_mismatch_details(ground_truth_path: str, predicted_path: str) -> pd.DataFrame:
    """
    Compare files and return a DataFrame containing only mismatched records.
    Columns: cik, year, Actual_{col}, Pred_{col}
    """
    gt_path = Path(ground_truth_path)
    pred_path = Path(predicted_path)

    if not gt_path.exists():
        raise FileNotFoundError(f"Missing: {gt_path}")
    if not pred_path.exists():
        raise FileNotFoundError(f"Missing: {pred_path}")

    # Load data
    gt = pd.read_csv(gt_path)
    pred = pd.read_csv(pred_path)

    # Filter columns
    gt_cols = ["cik", "year"] + COLUMNS_TO_VALIDATE
    pred_cols = ["cik", "year"] + COLUMNS_TO_VALIDATE

    gt = gt[gt_cols].copy()
    pred = pred[pred_cols].copy()

    # Merge
    merged = gt.merge(pred, on=["cik", "year"], suffixes=("_actual", "_pred"))

    # Track mismatches
    mismatch_indices = set()

    print("🔍 Analyzing mismatches...")

    for col in COLUMNS_TO_VALIDATE:
        actual_col = f"{col}_actual"
        pred_col = f"{col}_pred"

        # STRICT Type Casting (Handle 1 vs 1.0 vs "1")
        s_actual = (
            pd.to_numeric(merged[actual_col], errors="coerce").fillna(-1).astype(int)
        )
        s_pred = pd.to_numeric(merged[pred_col], errors="coerce").fillna(-1).astype(int)

        # Update DataFrame with clean integers for export
        merged[actual_col] = s_actual
        merged[pred_col] = s_pred

        # Find differences
        diff_mask = s_actual != s_pred
        if diff_mask.sum() > 0:
            print(f"   Found {diff_mask.sum()} mismatches in '{col}'")
            mismatch_indices.update(merged[diff_mask].index.tolist())

    if not mismatch_indices:
        return pd.DataFrame()

    # Return only the mismatched rows
    result_df = merged.loc[list(mismatch_indices)].copy()

    # Rename columns for clarity in Excel (e.g., cp_user_actual -> Actual_cp_user)
    rename_map = {}
    for col in COLUMNS_TO_VALIDATE:
        rename_map[f"{col}_actual"] = f"Actual_{col}"
        rename_map[f"{col}_pred"] = f"Pred_{col}"

    result_df = result_df.rename(columns=rename_map)
    return result_df


# =============================================================================
# 2. FETCH DATABASE METADATA
# =============================================================================


def fetch_db_data(db_path: str, mismatch_keys: pd.DataFrame) -> pd.DataFrame:
    """
    Fetch URL and Categories for the specific CIK/Year pairs found in mismatches.
    """
    conn = sqlite3.connect(db_path, timeout=60)
    query = """
        SELECT rd.url, rd.cik, rd.year, cat.categories
        FROM report_data rd
        JOIN category cat ON rd.url = cat.url
    """
    # Read entire mapping (usually faster than 1000s of SELECTs for medium datasets)
    # If DB is huge, filtering via SQL IN clause would be better, but this is safe for <1M rows
    db_df = pd.read_sql_query(query, conn)
    conn.close()

    # Convert DB types to match mismatch_keys for merging
    db_df["cik"] = pd.to_numeric(db_df["cik"], errors="coerce").fillna(0).astype(int)
    db_df["year"] = pd.to_numeric(db_df["year"], errors="coerce").fillna(0).astype(int)

    # Filter DB results to only those in our mismatch list
    # Inner join will keep URL/Categories and duplicate the Mismatch Values if multiple URLs exist
    keys = mismatch_keys[["cik", "year"]].drop_duplicates()
    filtered_db = db_df.merge(keys, on=["cik", "year"], how="inner")

    return filtered_db


# =============================================================================
# 3. EXCEL FORMATTING & EXPORT
# =============================================================================


def format_excel(ws):
    """Apply styling to the header and cells."""
    # Header Style
    header_fill = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Column Widths
    ws.column_dimensions["A"].width = 60  # URL
    ws.column_dimensions["B"].width = 12  # CIK
    ws.column_dimensions["C"].width = 8  # Year

    # Auto-width for value columns (D, E, etc.)
    col_idx = 4
    while True:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val == "Categories":
            ws.column_dimensions[col_letter].width = 50
            break
        elif header_val is None:
            break
        else:
            ws.column_dimensions[col_letter].width = 15  # Actual/Pred columns
        col_idx += 1

    # Row Styling
    light_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = light_fill
            cell.border = border

            # Alignments
            if cell.column == 1:  # URL
                cell.alignment = Alignment(horizontal="left")
            elif cell.column == len(row):  # Categories (last col)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            else:  # CIK, Year, Values
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


def export_mismatches(db_path: str, excel_path: Optional[str] = None):
    if excel_path is None:
        excel_path = "mismatches_debug.xlsx"

    print(f"{'='*70}\nEXPORTING MISMATCHES (WITH DEBUG VALUES)\n{'='*70}")

    # 1. Get Mismatch Data (CIK, Year, Values)
    try:
        df_mismatches = get_mismatch_details(GROUND_TRUTH_FILE, PREDICTED_FILE)
    except Exception as e:
        print(f"❌ Error: {e}")
        return

    if df_mismatches.empty:
        print("✅ No mismatches found!")
        return

    print(f"   Found {len(df_mismatches)} mismatched filing years.")

    # 2. Get DB Data (URL, Categories)
    print("🔄 Fetching URLs and Categories from DB...")
    df_db = fetch_db_data(db_path, df_mismatches)

    # 3. Merge Metadata with Mismatch Values
    final_df = df_db.merge(df_mismatches, on=["cik", "year"], how="left")

    # 4. Clean up JSON Categories for CSV/Excel
    print("📝 formatting categories...")

    def clean_json(x):
        try:
            return ", ".join(json.loads(x)) if x else ""
        except:
            return "JSON Error"

    final_df["categories"] = final_df["categories"].apply(clean_json)

    # 5. Reorder Columns
    # Standard: URL, CIK, Year, [Actual_X, Pred_X...], Categories
    value_cols = [c for c in df_mismatches.columns if c not in ["cik", "year"]]
    cols_order = ["url", "cik", "year"] + value_cols + ["categories"]
    final_df = final_df[cols_order]

    # 6. Write to Excel
    print(f"💾 Saving to {excel_path}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mismatches"

    # Write headers
    ws.append(list(final_df.columns))

    # Write rows
    for row in tqdm(
        final_df.itertuples(index=False), total=len(final_df), desc="Writing rows"
    ):
        ws.append(list(row))

    format_excel(ws)
    wb.save(excel_path)
    print(f"\n✅ Done! File saved: {excel_path}")


if __name__ == "__main__":
    import sys

    db_name = sys.argv[1] if len(sys.argv) > 1 else "classified_data.db"
    export_mismatches(db_name)
